import re
from openpyxl import load_workbook


__alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
              'V', 'W', 'X', 'Y', 'Z']


def convertLetterToColumnNum(columnName):
    converted = 0
    columnCount = len(columnName)

    base = len(__alphabet)
    base_exponent = 1
    while len(columnName) != 0:
        # strip the right-most digit, convert to index
        digit = columnName[-1:]
        columnName = columnName[:-1]
        digit = __alphabet.index(digit)

        # add the value it represents to the total, increment base_exponent
        converted = converted + digit * base_exponent
        base_exponent = base_exponent * base

        # add the offset for having passed all the n-width columns
        if len(columnName) != 0:
            converted = converted + base_exponent

    return converted + 1

def pars(Path_to_file:str, ):

    global dict_, list_resources, dict_resources
    rb = load_workbook(Path_to_file)
    sheet = rb[rb.sheetnames[0]]
    flag = 0
    first = 0
    data = []
    for i in range(8, sheet.max_row):
        cell = sheet.cell(row=i, column=1)
        if type(cell.value)!=type(None):
            if re.findall(r'Этап', cell.value):
                if flag:
                    flag = 0
                    dict_ = {'Stage': cell.value}

                else:
                    flag = 1
                    first = 1
                    dict_ = {'Stage': cell.value}

                continue
            if re.findall(r'ООО', cell.value):
                dict_OOO = {'ООО': cell.value}
                dict_resources = {}
                list_resources = []
            if first:
                merg_cells = sheet.merged_cells
                if cell.coordinate in merg_cells:
                    cell_merg_letters = [o for o in merg_cells if cell.coordinate in o]
                    # cell_merg_letters[0].coord
                    LETTERS = re.findall(r'[A-Z]', cell_merg_letters[0].coord)
                    first_col = convertLetterToColumnNum(str(LETTERS[0]))
                    second_col = convertLetterToColumnNum(str(LETTERS[1]))
                    span = abs(second_col-first_col) + 1
                    print(span)
                    if span == 3:
                        dict_resources = {'resource': cell.value, 'value': sheet.cell(row=i, column=14).value}
                        data.append({**dict_OOO, **dict_, **dict_resources})
    return data

